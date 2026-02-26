"""
AlphaAgent — Multi-Agent Orchestrator
Routes user queries through specialized financial agent personas with tool-calling skills.

Architecture:
    User Query → Orchestrator
        → Step 1: Detect query type (RAG, memo, compliance, comparative)
        → Step 2: Retrieve relevant documents via embedding search
        → Step 3: Load structured metrics from XLSX if ticker detected
        → Step 4: Optionally run compliance check
        → Step 5: Synthesize final response via Nemotron LLM with citations

All processing uses NVIDIA NIM microservices running locally in the Azure VNet.
All data resides on Azure NetApp Files — zero data movement.
"""

from __future__ import annotations

import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .config import Settings, get_settings
from .nim_client import chat_completion
from .indexer import load_index, query_index
from .prompts import (
    ORCHESTRATOR_PROMPT,
    SEC_ANALYST_PROMPT,
    QUANT_ANALYST_PROMPT,
    COMPLIANCE_OFFICER_PROMPT,
    SUMMARIZATION_PROMPT,
)
from .skills.financial_math import calculate_yoy_variance, calculate_margin, calculate_leverage
from .skills.compliance_checker import run_compliance_check, format_compliance_report


# Known tickers in the synthetic dataset
TICKERS = ["ALPH", "BETA", "GAMM"]


@dataclass
class AgentStep:
    """Records a single step in the agent's chain of thought."""
    agent: str
    action: str
    input_summary: str
    output_summary: str
    duration_ms: int = 0


@dataclass
class AgentResult:
    """Complete result from a multi-agent query."""
    answer: str = ""
    trace: List[AgentStep] = field(default_factory=list)
    citations: List[Dict] = field(default_factory=list)
    compliance: Optional[Dict] = None
    math_results: List[Dict] = field(default_factory=list)
    total_ms: int = 0


def detect_tickers(text: str) -> List[str]:
    """Detect company tickers in user query."""
    up = text.upper()
    found = []
    for t in TICKERS:
        if re.search(rf"\b{re.escape(t)}\b", up):
            found.append(t)
    return found


def detect_query_type(text: str) -> str:
    """Determine the type of query for agent routing."""
    lower = text.lower()
    if any(w in lower for w in ["memo", "investment brief", "investment memo"]):
        return "memo"
    if any(w in lower for w in ["compliance", "policy", "regulation", "surveillance", "audit"]):
        return "compliance"
    if any(w in lower for w in ["compare", "versus", "vs", "across"]):
        return "comparative"
    if any(w in lower for w in ["calculate", "variance", "yoy", "margin", "ratio"]):
        return "math"
    return "rag"


def load_metrics(data_root: str, ticker: str) -> Dict[str, str]:
    """
    Read structured financial metrics from the XLSX spreadsheet on ANF.
    Returns key-value pairs for the given ticker.
    """
    xlsx = Path(data_root) / "spreadsheets" / f"{ticker}_Key_Metrics.xlsx"
    if not xlsx.exists():
        return {}

    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    ws = wb["Key Metrics"] if "Key Metrics" in wb.sheetnames else wb.worksheets[0]

    metrics: Dict[str, str] = {}
    start_row = None
    for r in range(1, 50):
        v = ws.cell(row=r, column=1).value
        if str(v).strip().lower() == "metric":
            start_row = r + 1
            break

    if not start_row:
        return metrics

    r = start_row
    while r < start_row + 200:
        m = ws.cell(row=r, column=1).value
        if m is None:
            break
        val = ws.cell(row=r, column=2).value
        unit = ws.cell(row=r, column=3).value or ""
        note = ws.cell(row=r, column=4).value or ""
        metrics[str(m)] = f"{val} {unit} ({note})" if note else f"{val} {unit}"
        r += 1
    return metrics


def load_metrics_numeric(data_root: str, ticker: str) -> Dict[str, float]:
    """Load numeric metric values for calculations."""
    xlsx = Path(data_root) / "spreadsheets" / f"{ticker}_Key_Metrics.xlsx"
    if not xlsx.exists():
        return {}

    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    ws = wb["Key Metrics"] if "Key Metrics" in wb.sheetnames else wb.worksheets[0]

    metrics: Dict[str, float] = {}
    start_row = None
    for r in range(1, 50):
        v = ws.cell(row=r, column=1).value
        if str(v).strip().lower() == "metric":
            start_row = r + 1
            break

    if not start_row:
        return metrics

    r = start_row
    while r < start_row + 200:
        m = ws.cell(row=r, column=1).value
        if m is None:
            break
        val = ws.cell(row=r, column=2).value
        try:
            metrics[str(m)] = float(val)
        except (ValueError, TypeError):
            pass
        r += 1
    return metrics


def run_query(
    question: str,
    settings: Optional[Settings] = None,
    records: Optional[List] = None,
    matrix=None,
) -> AgentResult:
    """
    Execute a multi-agent financial research query.

    This is the main entry point for the AlphaAgent orchestrator.
    It routes through specialized agents based on query type.
    """
    s = settings or get_settings()
    result = AgentResult()
    start = time.time()

    # ── Step 1: Orchestrator — Classify and plan ──
    t0 = time.time()
    query_type = detect_query_type(question)
    tickers = detect_tickers(question)
    result.trace.append(AgentStep(
        agent="🧠 Orchestrator",
        action="Classify & Plan",
        input_summary=question[:100],
        output_summary=f"Type: {query_type} | Tickers: {tickers or 'none detected'}",
        duration_ms=int((time.time() - t0) * 1000),
    ))

    # ── Step 2: SEC Analyst — Retrieve relevant documents from ANF ──
    t0 = time.time()
    if records is None or matrix is None:
        records, matrix = load_index(s.index_root)

    search_query = f"{' '.join(tickers)} {question}" if tickers else question
    hits = query_index(
        query=search_query,
        embed_base_url=s.embed_base_url,
        embed_model=s.embed_model,
        records=records,
        matrix=matrix,
        top_k=s.top_k,
    )
    result.citations = hits
    result.trace.append(AgentStep(
        agent="🔍 SEC Research Analyst",
        action="Document Retrieval (ANF → NIM Embeddings → Cosine Search)",
        input_summary=search_query[:80],
        output_summary=f"Retrieved {len(hits)} chunks from {len(set(h['doc_id'] for h in hits))} documents",
        duration_ms=int((time.time() - t0) * 1000),
    ))

    # ── Step 3: Quant Analyst — Load metrics + run calculations ──
    all_metrics: Dict[str, Dict] = {}
    all_metrics_numeric: Dict[str, Dict] = {}

    for ticker in tickers:
        t0 = time.time()
        metrics = load_metrics(s.data_root, ticker)
        metrics_num = load_metrics_numeric(s.data_root, ticker)
        all_metrics[ticker] = metrics
        all_metrics_numeric[ticker] = metrics_num

        # Run financial calculations if data available
        if "CapEx_Current" in metrics_num and "CapEx_Prior" in metrics_num:
            yoy = calculate_yoy_variance(metrics_num["CapEx_Current"], metrics_num["CapEx_Prior"])
            result.math_results.append({"ticker": ticker, "calculation": "CapEx YoY Variance", **yoy})

        if "EBITDA_TTM" in metrics_num and "Revenue_TTM" in metrics_num:
            margin = calculate_margin(metrics_num["EBITDA_TTM"], metrics_num["Revenue_TTM"])
            result.math_results.append({"ticker": ticker, "calculation": "EBITDA Margin", **margin})

        if "NetDebt_to_EBITDA" in metrics_num:
            lev = calculate_leverage(
                metrics_num.get("NetDebt_to_EBITDA", 0) * metrics_num.get("EBITDA_TTM", 1),
                metrics_num.get("EBITDA_TTM", 1),
            )
            result.math_results.append({"ticker": ticker, "calculation": "Leverage Ratio", **lev})

        if metrics:
            result.trace.append(AgentStep(
                agent="📊 Quant Analyst",
                action=f"Load Metrics + Calculate ({ticker})",
                input_summary=f"XLSX from ANF: {ticker}_Key_Metrics.xlsx",
                output_summary=f"Loaded {len(metrics)} metrics, ran {len(result.math_results)} calculations",
                duration_ms=int((time.time() - t0) * 1000),
            ))

    # ── Step 4: Compliance Officer — Policy check (if relevant) ──
    if query_type in ("compliance", "memo") and tickers:
        for ticker in tickers:
            t0 = time.time()
            metrics_num = all_metrics_numeric.get(ticker, {})
            check_metrics = {}

            if "CapEx_Current" in metrics_num and "CapEx_Prior" in metrics_num:
                yoy_pct = ((metrics_num["CapEx_Current"] - metrics_num["CapEx_Prior"]) / metrics_num["CapEx_Prior"]) * 100
                check_metrics["capex_yoy_pct"] = yoy_pct

            if "NetDebt_to_EBITDA" in metrics_num:
                check_metrics["leverage_ratio"] = metrics_num["NetDebt_to_EBITDA"]

            if "VaR_99_1d" in metrics_num:
                check_metrics["var_99_usd_m"] = metrics_num["VaR_99_1d"]

            if check_metrics:
                compliance_result = run_compliance_check(check_metrics, ticker=ticker)
                result.compliance = compliance_result
                result.trace.append(AgentStep(
                    agent="✅ Compliance Officer",
                    action=f"Policy Threshold Check ({ticker})",
                    input_summary=f"Checking {len(check_metrics)} metrics against internal policies",
                    output_summary=f"Status: {compliance_result['overall_status']} — {compliance_result['flags']} flags",
                    duration_ms=int((time.time() - t0) * 1000),
                ))

    # ── Step 5: Synthesize — Generate final response via Nemotron ──
    t0 = time.time()

    # Build context from retrieved documents
    context_lines = []
    for h in hits:
        context_lines.append(f"[Source: {h['doc_id']} | score={h['score']:.3f}] {h['text']}")

    # Build metrics context
    metrics_lines = []
    for ticker, metrics in all_metrics.items():
        metrics_lines.append(f"\n--- {ticker} Key Metrics (from XLSX on ANF) ---")
        for k, v in metrics.items():
            metrics_lines.append(f"  {k}: {v}")

    # Build math results context
    math_lines = []
    for mr in result.math_results:
        math_lines.append(f"  [{mr.get('ticker', '')}] {mr.get('calculation', '')}: {mr.get('result', '')}")
        if "interpretation" in mr:
            math_lines.append(f"    → {mr['interpretation']}")

    # Build compliance context
    compliance_text = ""
    if result.compliance:
        compliance_text = f"\n--- Compliance Assessment ---\n{format_compliance_report(result.compliance)}"

    # Select the appropriate system prompt based on query type
    if query_type == "memo":
        system_prompt = SUMMARIZATION_PROMPT
    elif query_type == "compliance":
        system_prompt = COMPLIANCE_OFFICER_PROMPT
    elif query_type == "math":
        system_prompt = QUANT_ANALYST_PROMPT
    else:
        system_prompt = ORCHESTRATOR_PROMPT

    user_message = (
        f"User Question:\n{question}\n\n"
        f"RETRIEVED DOCUMENTS (from Azure NetApp Files via NIM embedding search):\n"
        + "\n\n".join(context_lines)
        + ("\n\nSTRUCTURED METRICS (from XLSX on ANF):\n" + "\n".join(metrics_lines) if metrics_lines else "")
        + ("\n\nCALCULATION RESULTS (deterministic, not generated):\n" + "\n".join(math_lines) if math_lines else "")
        + (compliance_text if compliance_text else "")
    )

    answer = chat_completion(
        llm_base_url=s.llm_base_url,
        model=s.llm_model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_message},
        ],
        max_tokens=s.llm_max_tokens,
        temperature=s.llm_temperature,
    )

    result.answer = answer
    result.trace.append(AgentStep(
        agent="✍️ Nemotron LLM (Synthesis)",
        action="Generate Response with Citations",
        input_summary=f"Context: {len(hits)} doc chunks + {len(all_metrics)} ticker metrics",
        output_summary=f"Generated {len(answer)} chars",
        duration_ms=int((time.time() - t0) * 1000),
    ))

    result.total_ms = int((time.time() - start) * 1000)
    return result
