"""
AlphaAgent — Embedding Indexer
Builds and queries a flat JSONL + numpy embedding index on Azure NetApp Files.
No vector DB required — keeps the demo simple and reliable.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import numpy as np
from pypdf import PdfReader
from openpyxl import load_workbook

from .nim_client import embed_texts


@dataclass(frozen=True)
class Chunk:
    doc_id: str
    chunk_id: int
    text: str
    source_path: str


def _iter_files(data_root: Path) -> Iterable[Path]:
    """Yield all PDF and XLSX files under data_root."""
    for p in sorted(data_root.rglob("*")):
        if p.is_file() and p.suffix.lower() in [".pdf", ".xlsx"]:
            yield p


def _read_pdf(path: Path) -> str:
    """Extract text from a PDF file."""
    reader = PdfReader(str(path))
    texts: List[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        texts.append(t)
    return "\n".join(texts)


def _read_xlsx(path: Path) -> str:
    """Extract text from an XLSX file (all sheets)."""
    wb = load_workbook(str(path), read_only=True, data_only=True)
    texts: List[str] = []
    for ws in wb.worksheets:
        texts.append(f"[SHEET] {ws.title}")
        for row in ws.iter_rows(values_only=True):
            vals = [str(v) for v in row if v is not None and str(v).strip() != ""]
            if vals:
                texts.append(" | ".join(vals))
    return "\n".join(texts)


def _chunk_text(text: str, chunk_chars: int, overlap: int) -> List[str]:
    """Split text into overlapping chunks."""
    text = " ".join(text.split())  # Normalize whitespace
    if not text:
        return []
    chunks: List[str] = []
    i = 0
    while i < len(text):
        chunk = text[i : i + chunk_chars]
        chunks.append(chunk)
        if i + chunk_chars >= len(text):
            break
        i += max(1, chunk_chars - overlap)
    return chunks


def build_index(
    *,
    data_root: str,
    index_root: str,
    embed_base_url: str,
    embed_model: str,
    chunk_chars: int,
    chunk_overlap: int,
    batch_size: int = 16,
) -> Path:
    """
    Build an embedding index from documents on ANF.

    Creates:
      <index_root>/index.jsonl  — one JSON record per chunk (with embedding)
      <index_root>/manifest.json — metadata about the index
    """
    data_path = Path(data_root)
    index_path = Path(index_root)
    index_path.mkdir(parents=True, exist_ok=True)

    index_file = index_path / "index.jsonl"
    manifest_file = index_path / "manifest.json"

    # Idempotent check
    if index_file.exists():
        print(f"  Index already exists at {index_file}, skipping rebuild.")
        return index_file

    # Read and chunk all documents
    chunks: List[Chunk] = []
    for f in _iter_files(data_path):
        print(f"  Indexing: {f.name}")
        if f.suffix.lower() == ".pdf":
            raw = _read_pdf(f)
        else:
            raw = _read_xlsx(f)

        doc_id = f.name
        for idx, ch in enumerate(_chunk_text(raw, chunk_chars, chunk_overlap)):
            chunks.append(Chunk(doc_id=doc_id, chunk_id=idx, text=ch, source_path=str(f)))

    if not chunks:
        print("  ⚠️  No documents found to index.")
        return index_file

    # Generate embeddings in batches via NIM
    print(f"  Embedding {len(chunks)} chunks in batches of {batch_size}...")
    embeddings: List[List[float]] = []
    for i in range(0, len(chunks), batch_size):
        batch = chunks[i : i + batch_size]
        vecs = embed_texts(embed_base_url, embed_model, [c.text for c in batch], input_type="passage")
        embeddings.extend(vecs)
        if (i // batch_size) % 5 == 0:
            print(f"    Batch {i // batch_size + 1}/{(len(chunks) + batch_size - 1) // batch_size}")

    # Write JSONL index
    with index_file.open("w", encoding="utf-8") as f:
        for c, e in zip(chunks, embeddings):
            rec = {
                "doc_id": c.doc_id,
                "chunk_id": c.chunk_id,
                "source_path": c.source_path,
                "text": c.text,
                "embedding": e,
            }
            f.write(json.dumps(rec) + "\n")

    # Write manifest
    manifest = {
        "num_chunks": len(chunks),
        "num_documents": len(set(c.doc_id for c in chunks)),
        "data_root": str(data_path),
        "index_file": str(index_file),
        "embed_model": embed_model,
        "chunk_chars": chunk_chars,
        "chunk_overlap": chunk_overlap,
    }
    manifest_file.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    print(f"  ✅ Index built: {len(chunks)} chunks from {manifest['num_documents']} documents.")
    return index_file


def load_index(index_root: str) -> Tuple[List[Dict], np.ndarray]:
    """Load the JSONL index into memory and return records + normalized embedding matrix."""
    index_file = Path(index_root) / "index.jsonl"
    if not index_file.exists():
        raise FileNotFoundError(f"Index not found: {index_file}")

    recs: List[Dict] = []
    vecs: List[List[float]] = []
    for line in index_file.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        recs.append(rec)
        vecs.append(rec["embedding"])

    mat = np.array(vecs, dtype=np.float32)
    # L2+normalize for cosine similarity via dot product
    norms = np.linalg.norm(mat, axis=1, keepdims=True) + 1e-12
    mat = mat / norms

    return recs, mat


def query_index(
    *,
    query: str,
    embed_base_url: str,
    embed_model: str,
    records: List[Dict],
    matrix: np.ndarray,
    top_k: int = 5,
) -> List[Dict]:
    """Semantic search: embed the query and return top-k most similar chunks."""
    qvec = embed_texts(embed_base_url, embed_model, [query], input_type="query")[0]
    q = np.array(qvec, dtype=np.float32)
    q = q / (np.linalg.norm(q) + 1e-12)

    scores = matrix @ q
    idxs = np.argsort(-scores)[:top_k]

    out: List[Dict] = []
    for i in idxs:
        rec = dict(records[int(i)])
        rec.pop("embedding", None)  # Don't return raw embeddings
        rec["score"] = float(scores[int(i)])
        out.append(rec)
    return out
